import openpyxl
import sys

wb = openpyxl.load_workbook(r'd:\桌面\system-inspection-main\巡检ip汇总.xlsx')
print("Sheets:", wb.sheetnames)

for sname in wb.sheetnames:
    ws = wb[sname]
    print("\n=== Sheet: {} (rows={}, cols={}) ===".format(sname, ws.max_row, ws.max_column))
    for row in ws.iter_rows(min_row=1, max_row=min(ws.max_row, 80), values_only=True):
        print(list(row))

sys.exit(0)
