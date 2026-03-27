"""
从 巡检ip汇总.xlsx 中提取所有有效 IP 并生成 config.json 配置。
运行方式：python import_excel_config.py
"""
import openpyxl
import json
import os
import re

EXCEL_FILE = r"d:\桌面\system-inspection-main\巡检ip汇总.xlsx"
CONFIG_FILE = r"d:\桌面\system-inspection-main\data\config.json"

IP_PATTERN = re.compile(r"^\d{1,3}\.\d{1,3}\.\d{1,3}\.\d{1,3}$")


def is_valid_ip(val):
    if not val or not isinstance(val, str):
        return False
    return bool(IP_PATTERN.match(val.strip()))


def find_ip_column(ws, header_row=1):
    """自动查找含 'IP' 关键字的列索引"""
    for row in ws.iter_rows(min_row=1, max_row=min(3, ws.max_row), values_only=False):
        for cell in row:
            val = str(cell.value or "")
            if "IP" in val.upper() or "地址" in val:
                return cell.column - 1  # 0-indexed
    return None


def extract_hosts_generic(ws, sheet_name):
    """
    通用提取逻辑：扫描所有行和列，
    自动识别 IP 列，提取设备名称和 IP。
    """
    hosts = []

    # 先找表头行和各列位置
    ip_col = None
    name_cols = []  # 可能用于名称的列（区域、用途、编号等）
    header_row_idx = None

    for row_idx, row in enumerate(ws.iter_rows(min_row=1, max_row=min(5, ws.max_row), values_only=True)):
        for col_idx, val in enumerate(row):
            val_str = str(val or "")
            if "IP" in val_str.upper() or val_str == "地址":
                ip_col = col_idx
                header_row_idx = row_idx + 1  # 1-indexed
                break
        if ip_col is not None:
            break

    if ip_col is None:
        # 没找到IP列表头，尝试扫描所有单元格找IP
        for row in ws.iter_rows(min_row=1, values_only=True):
            for col_idx, val in enumerate(row):
                if is_valid_ip(str(val or "").strip()):
                    ip_col = col_idx
                    header_row_idx = 1
                    break
            if ip_col is not None:
                break

    if ip_col is None:
        return hosts

    # 确定名称列：IP列之前的所有列
    name_cols = list(range(0, ip_col))

    # 遍历数据行
    last_names = [""] * len(name_cols)
    start_row = (header_row_idx or 1) + 1

    for row in ws.iter_rows(min_row=start_row, values_only=True):
        if len(row) <= ip_col:
            continue

        ip = str(row[ip_col] or "").strip()
        if not is_valid_ip(ip):
            continue

        # 构建名称
        parts = []
        for i, nc in enumerate(name_cols):
            if nc < len(row) and row[nc]:
                val = str(row[nc]).replace("\n", "").strip()
                last_names[i] = val
            if last_names[i]:
                parts.append(last_names[i])

        # 去重复词
        unique_parts = []
        for p in parts:
            if p not in unique_parts:
                unique_parts.append(p)

        name = "-".join(unique_parts) if unique_parts else sheet_name
        hosts.append({"name": name, "host": ip})

    return hosts


def main():
    if not os.path.exists(EXCEL_FILE):
        print("错误：找不到文件 {}".format(EXCEL_FILE))
        return

    wb = openpyxl.load_workbook(EXCEL_FILE)
    print("工作表: {}".format(wb.sheetnames))

    all_hosts = []

    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        hosts = extract_hosts_generic(ws, sheet_name)
        if hosts:
            print("[{}] 提取: {} 个设备".format(sheet_name, len(hosts)))
            all_hosts.extend(hosts)
        else:
            print("[{}] 无有效 IP，跳过".format(sheet_name))

    # 去重（按 IP 去重）
    seen_ips = set()
    unique_hosts = []
    for h in all_hosts:
        if h["host"] not in seen_ips:
            seen_ips.add(h["host"])
            unique_hosts.append(h)
    print("\n去重后: {} 个唯一 IP".format(len(unique_hosts)))

    # 读取现有配置（保留告警设置）
    config = {
        "URL_CHECKS": [],
        "ORACLE_DBS": [],
        "MYSQL_DBS": [],
        "HOSTS": [],
        "TCP_PORTS": [],
        "TELNET_CHECKS": [],
        "ALERT_CONFIG": {"url": "", "key": ""}
    }

    if os.path.exists(CONFIG_FILE):
        with open(CONFIG_FILE, "r", encoding="utf-8") as f:
            old_config = json.load(f)
            config["ALERT_CONFIG"] = old_config.get("ALERT_CONFIG", config["ALERT_CONFIG"])
            for k in ["URL_CHECKS", "ORACLE_DBS", "MYSQL_DBS", "TCP_PORTS", "TELNET_CHECKS"]:
                config[k] = old_config.get(k, [])

    config["HOSTS"] = unique_hosts

    os.makedirs(os.path.dirname(CONFIG_FILE), exist_ok=True)
    with open(CONFIG_FILE, "w", encoding="utf-8") as f:
        json.dump(config, f, ensure_ascii=False, indent=4)

    print("\n配置已保存到: {}".format(CONFIG_FILE))
    print("共导入 {} 个主机巡检项".format(len(unique_hosts)))

    print("\n--- 导入预览 ---")
    for i, h in enumerate(unique_hosts, 1):
        print("  {}. {} -> {}".format(i, h["name"], h["host"]))


if __name__ == "__main__":
    main()
